import pywhatkit
from openpyxl import load_workbook

def anindaMesajGonder(dosyaAdi, calismaSayfasi, isimSutunu, telefonSutunu, urunSutunu, telefonKodu, beklemeSuresi, sekmeKapatma, sekmeKapatmaOncesiBeklemeSuresi, kaldigiYerden):
    if kaldigiYerden:
        wb = load_workbook(dosyaAdi)
        sheet = wb[calismaSayfasi]
        sheet = wb.active

        isimSoyisimListesi = []
        telefonListesi = []
        urunListesi = []

        for cell in sheet[isimSutunu]:
            isimSoyisimListesi.append(str(cell.value))
        isimSoyisimListesi.remove('İsim Soyisim')

        for cell in sheet[telefonSutunu]:
            tempVal = str(cell.value)
            if (tempVal != None):
                tempVal = telefonKodu + tempVal
            telefonListesi.append(tempVal)
        telefonListesi.remove(telefonKodu + 'Telefon')

        for cell in sheet[urunSutunu]:
            urunListesi.append(str(cell.value))
        urunListesi.remove('Ürün')

        size = len(telefonListesi)

        onMesaj = "Gönderilecek {0} mesaj ve {1} kişi bulundu".format(size, size)
        print(onMesaj)

        dosya = open("WhatsAppMessageLog.txt", "r", encoding="utf-8")

        satirList = []

        for satir in dosya:
            satirList.append(satir)

        size = len(satirList)
        sonKelime = satirList[size - 4]
        sonKaldigiYer = [int(i) for i in sonKelime.split() if i.isdigit()]
        print(sonKaldigiYer)

        for i in range(size):
            if(i >= sonKaldigiYer[0] + 1):
                kampanyaMesaji = 'Sayın ' + isimSoyisimListesi[i] + ', \nsatın almış olduğunuz ' + urunListesi[i] + ' ürünümüzde kampanya başlamıştır !\n  \nSize Özel Kampanya Kodunuz: kampanya10\n  \nKeyifli Alışverişler Dileriz ! \nmagaza.com'
                print(telefonListesi[i] + kampanyaMesaji)
                pywhatkit.sendwhatmsg_instantly(telefonListesi[i], kampanyaMesaji, beklemeSuresi, sekmeKapatma, sekmeKapatmaOncesiBeklemeSuresi, i)
    else:
        wb = load_workbook(dosyaAdi)
        sheet = wb[calismaSayfasi]
        sheet = wb.active

        isimSoyisimListesi = []
        telefonListesi = []
        urunListesi = []

        for cell in sheet[isimSutunu]:
            isimSoyisimListesi.append(str(cell.value))
        isimSoyisimListesi.remove('İsim Soyisim')

        for cell in sheet[telefonSutunu]:
            tempVal = str(cell.value)
            if (tempVal != None):
                tempVal = telefonKodu + tempVal
            telefonListesi.append(tempVal)
        telefonListesi.remove(telefonKodu + 'Telefon')

        for cell in sheet[urunSutunu]:
            urunListesi.append(str(cell.value))
        urunListesi.remove('Ürün')

        size = len(telefonListesi)

        onMesaj = "Gönderilecek {0} mesaj ve {1} kişi bulundu".format(size, size)
        print(onMesaj)

        for i in range(size):
            kampanyaMesaji = 'Sayın ' + isimSoyisimListesi[i] + ', \nsatın almış olduğunuz ' + urunListesi[i] + ' ürünümüzde kampanya başlamıştır !\n  \nSize Özel Kampanya Kodunuz: kampanya10\n  \nKeyifli Alışverişler Dileriz ! \nmagaza.com'
            print(telefonListesi[i] + kampanyaMesaji)
            pywhatkit.sendwhatmsg_instantly(telefonListesi[i], kampanyaMesaji, beklemeSuresi, sekmeKapatma, sekmeKapatmaOncesiBeklemeSuresi, i)

anindaMesajGonder("data.xlsx", "Sayfa1", 'A', 'D', 'I', '+90', 5, True, 4, False)